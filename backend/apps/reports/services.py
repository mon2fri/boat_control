from __future__ import annotations

import csv
import html
import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_FILTER_OP_LABELS: dict[str, str] = {
    "equals": "equals",
    "not_equals": "not equal to",
    "contains": "contains",
    "not_contains": "does not contain",
    "eq": "equals",
    "neq": "not equal to",
    "ncontains": "does not contain",
}
_EXCEL_MAX_ROWS = 1_048_576
_EXCEL_MAX_COLUMNS = 16_384


def _escape_html(value: Any) -> str:
    if value is None:
        return ""
    return html.escape(str(value))


def _sanitize_csv_value(value: Any) -> str:
    if value is None:
        return ""
    str_val = str(value)
    if str_val and str_val[0] in ("=", "+", "-", "@", "\t", "\r"):
        str_val = f"'{str_val}"
    return str_val


def _humanize_rule_logic(value: Any) -> str:
    text = str(value or "")
    replacements = (
        ("ncontains", "does not contain"),
        ("contains", "contains"),
        ("neq", "does not equal"),
        ("gte", "greater than or equal to"),
        ("lte", "less than or equal to"),
        ("gt", "greater than"),
        ("lt", "less than"),
        ("eq", "equals"),
    )
    for operator, wording in replacements:
        text = text.replace(f" {operator} ", f" {wording} ")
    return text


def _metric(label: str, value: Any) -> str:
    return (
        f"<div class='metric'><b>{_escape_html(value)}</b><span>{_escape_html(label)}</span></div>"
    )


def _detail_header(
    key_columns: list[str],
    extra_columns: list[str] | None = None,
    hide_comparison: bool = False,
) -> str:
    identity = "".join(f"<th>{_escape_html(column)}</th>" for column in key_columns)
    if not identity:
        identity = "<th>Row</th>"
    extras = "".join(f"<th>{_escape_html(column)}</th>" for column in extra_columns or [])
    comparison = (
        "<th>Column</th><th>In Baseline</th><th>In Comparison</th>" if not hide_comparison else ""
    )
    return f"<tr>{identity}{extras}{comparison}</tr>"


def _identity_cells(key_columns: list[str], key_values: dict[str, Any], row_index: Any) -> str:
    if key_columns:
        return "".join(
            f"<td>{_escape_html(key_values.get(column, '—'))}</td>" for column in key_columns
        )
    return f"<td>{_escape_html(row_index)}</td>"


def _format_filter_row(f: dict[str, Any]) -> str:
    op = _FILTER_OP_LABELS.get(f.get("operator", ""), f.get("operator", ""))
    vals = f.get("filter_values") or []
    if not vals and f.get("filter_value"):
        vals = [f["filter_value"]]
    quoted = " or ".join(f"'{v}'" for v in vals)
    return f"{f.get('column', '?')} {op} {quoted}"


def _format_created_at(iso_str: str | None) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%b %d, %Y %H:%M:%S")
    except (ValueError, TypeError):
        return iso_str


def _group_stat_card(stat: dict[str, Any]) -> str:
    col = _escape_html(stat.get("column", "?"))
    unique = _escape_html(stat.get("unique_count", 0))
    rows_html = ""
    for r in stat.get("rows", []):
        val = _escape_html(r.get("value", ""))
        ru = _escape_html(r.get("unique_count", 0))
        rows_html += f"<tr><td>{val}</td><td>{ru}</td></tr>"
    return (
        f"<details class='group-stats-card'>"
        f"<summary class='group-stats-summary'>"
        f"<span class='group-stats-column'>{col}</span>"
        f"<span class='group-stats-count'>Exception records: {unique}</span></summary>"
        f"<div class='group-stats-scroll'><table class='group-stats-table'>"
        f"<thead><tr><th>Value</th><th>Exception records</th></tr></thead>"
        f"<tbody>{rows_html}</tbody>"
        f"</table></div>"
        f"</details>"
    )


def _distribute_evenly(
    items: list[dict[str, Any]], max_per_row: int = 4
) -> list[list[dict[str, Any]]]:
    """Mirror the result page's balanced aggregation-card row layout."""
    if not items:
        return []
    if len(items) <= max_per_row:
        return [items]
    row_count = (len(items) + max_per_row - 1) // max_per_row
    base_size, remainder = divmod(len(items), row_count)
    rows: list[list[dict[str, Any]]] = []
    offset = 0
    for index in range(row_count):
        size = base_size + (1 if index < remainder else 0)
        rows.append(items[offset : offset + size])
        offset += size
    return rows


def _render_group_section(title: str, stats: list[dict[str, Any]]) -> str:
    if not stats:
        return ""
    rows = []
    for row in _distribute_evenly(stats):
        cards = "".join(_group_stat_card(stat) for stat in row)
        rows.append(
            f"<div class='group-stats-row group-stats-row--{min(len(row), 4)}'>{cards}</div>"
        )
    return (
        f"<div class='group-stats-panel' aria-label='{_escape_html(title)}'>{''.join(rows)}</div>"
    )


def _render_exception_rule_summary(validation: dict[str, Any]) -> str:
    violations_by_rule = validation.get("violations_by_rule") or {}
    summaries = validation.get("rule_summaries") or {}
    rule_ids = list(dict.fromkeys([*violations_by_rule, *summaries]))
    if not rule_ids:
        return (
            "<section class='card'><h2>Exception Rule Summary</h2>"
            "<p>No exception rules were selected.</p></section>"
        )

    rows = []
    row_counts = validation.get("violating_rows_by_rule") or {}
    for rule_id in rule_ids:
        summary = summaries.get(rule_id) or {}
        violations = violations_by_rule.get(rule_id) or []
        sample = violations[0] if violations else {}
        name = summary.get("name") or sample.get("rule_name") or rule_id
        count = row_counts.get(rule_id, len(violations))
        rows.append(
            "<tr>"
            f"<td><span class='rule-id'>{_escape_html(rule_id)}</span>"
            f"{_escape_html(name)}</td>"
            f"<td class='number-cell'>{_escape_html(count)}</td>"
            "</tr>"
        )
    return (
        "<section class='card'><h2>Exception Rule Summary</h2>"
        "<div class='table-scroll'><table class='result-table exception-rule-table'>"
        "<thead><tr><th>Rule name</th><th>Exception records</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table></div></section>"
    )


def _render_comparing_columns(result: dict[str, Any]) -> str:
    columns = result.get("target_columns")
    if not columns:
        columns = result.get("common_columns") or []
    tags = "".join(
        f"<li><span class='tag'>{_escape_html(column)}</span></li>" for column in columns
    )
    if not tags:
        return (
            "<div class='comparison-columns'><p class='comparison-columns-label'>"
            "Comparing columns</p><p class='section-logic'>No comparing columns.</p></div>"
        )
    return (
        "<div class='comparison-columns'><p class='comparison-columns-label'>"
        "Comparing columns</p><ul class='chip-list' aria-label='Comparing columns'>"
        f"{tags}</ul></div>"
    )


def _exception_table_columns(
    result: dict[str, Any], violations_by_rule: dict[str, list[dict[str, Any]]]
) -> list[str]:
    """Return configured exception columns, preserving old runs without configuration."""
    configured = result.get("exception_columns")
    if isinstance(configured, list):
        return [str(column) for column in configured]
    return list(
        dict.fromkeys(
            column
            for violations in violations_by_rule.values()
            for violation in violations
            for column in (violation.get("extra_values") or {})
        )
    )


def _render_exception_table(
    result: dict[str, Any], validation: dict[str, Any], key_columns: list[str]
) -> str:
    """Render the cross-rule table used by server-side HTML exports."""
    violations_by_rule = validation.get("violations_by_rule") or {}
    aggregation_columns = list(result.get("aggregation_columns") or [])
    aggregation_labels = result.get("aggregation_column_labels") or {}
    exception_columns = _exception_table_columns(result, violations_by_rule)
    headers = [
        *(key_columns or ["Row"]),
        "Rule Index",
        *(aggregation_labels.get(column, column) for column in aggregation_columns),
        *exception_columns,
    ]
    data_rows: list[list[str]] = []
    for rule_id, violations in violations_by_rule.items():
        for violation in violations:
            grouping = violation.get("grouping_values") or {}
            extras = violation.get("extra_values") or {}
            identity_values = (
                [
                    violation.get("key_columns", {}).get(column, "—")
                    for column in key_columns
                ]
                if key_columns
                else [violation.get("row_index", "")]
            )
            row_values = [
                *identity_values,
                rule_id,
                *(grouping.get(column, "—") for column in aggregation_columns),
                *(extras.get(column, "—") for column in exception_columns),
            ]
            data_rows.append([str(value) if value is not None else "—" for value in row_values])
    if not data_rows:
        return (
            "<section class='card' id='exception-table'><details>"
            "<summary>Exception Table (0 rows)</summary>"
            "<p>No exceptions found across any rule.</p></details></section>"
        )
    header_cells_list: list[str] = []
    for index, header in enumerate(headers):
        options = "".join(
            f'<option value="{_escape_html(value)}">{_escape_html(value)}</option>'
            for value in sorted({row[index] for row in data_rows})
        )
        header_cells_list.append(
            "<th>"
            f"<button type='button' class='exception-table-sort' data-exception-sort='{index}' "
            f"aria-label='Sort by {_escape_html(header)}'>{_escape_html(header)} ↕</button>"
            f"<select class='exception-table-filter' data-exception-filter='{index}' "
            f"aria-label='Filter {_escape_html(header)}'><option value=''>All</option>"
            f"{options}</select></th>"
        )
    header_cells = "".join(header_cells_list)
    body_rows = "".join(
        f"<tr>{''.join(f'<td>{_escape_html(value)}</td>' for value in row)}</tr>"
        for row in data_rows
    )
    return (
        "<section class='card' id='exception-table'><details>"
        f"<summary>Exception Table ({len(data_rows)} rows)</summary>"
        "<div class='table-scroll'><table class='result-table' data-exception-table><thead>"
        f"<tr>{header_cells}</tr></thead><tbody>{body_rows}</tbody></table></div>"
        "</details></section>"
    )


_EXCEPTION_TABLE_INTERACTIONS = """
<script type='text/javascript'>
document.addEventListener('DOMContentLoaded', function () {
  document.querySelectorAll('[data-exception-table]').forEach(function (table) {
    var body = table.tBodies[0];
    function applyFilters() {
      var filters = table.querySelectorAll('[data-exception-filter]');
      Array.from(body.rows).forEach(function (row) {
        row.hidden = Array.from(filters).some(function (filter) {
          return filter.value &&
            row.cells[Number(filter.dataset.exceptionFilter)].textContent !== filter.value;
        });
      });
    }
    table.querySelectorAll('[data-exception-filter]').forEach(function (filter) {
      filter.addEventListener('change', applyFilters);
    });
    table.querySelectorAll('[data-exception-sort]').forEach(function (button) {
      button.addEventListener('click', function () {
        var index = Number(button.dataset.exceptionSort);
        var direction = button.dataset.direction === 'asc' ? 'desc' : 'asc';
        table.querySelectorAll('[data-exception-sort]').forEach(function (item) {
          item.dataset.direction = item === button ? direction : '';
        });
        Array.from(body.rows).sort(function (left, right) {
          return left.cells[index].textContent.localeCompare(
            right.cells[index].textContent, undefined, { numeric: true, sensitivity: 'base' }
          ) * (direction === 'asc' ? 1 : -1);
        }).forEach(function (row) { body.appendChild(row); });
      });
    });
  });
});
</script>
"""


def export_html(result: dict[str, Any], report_name: str, created_at: str | None = None) -> str:
    comparison = result.get("comparison", {})
    validation = result.get("validation", {})
    key_columns = list(result.get("key_columns") or [])

    sections: list[str] = []

    sections.append("<!DOCTYPE html>")
    sections.append("<html><head>")
    sections.append(f"<title>{_escape_html(report_name)}</title>")
    sections.append("<style>")
    sections.append(":root { color-scheme: light; }")
    sections.append(
        "body { background:#f5f7fa; color:#1f2937; "
        "font-family:system-ui,sans-serif; margin:0; padding:24px; }"
    )
    sections.append("main { max-width:1200px; margin:0 auto; }")
    sections.append(
        ".report-header { display:flex; align-items:center; gap:12px; margin-bottom:12px; }"
    )
    sections.append(".report-header h1 { margin:0; font-size:1.35rem; }")
    sections.append(".run-time, .section-logic { color:#64748b; font-size:.85rem; }")
    sections.append(
        ".card { background:#fff; border:1px solid #dce2ea; "
        "border-radius:8px; padding:16px; margin-bottom:12px; }"
    )
    sections.append(
        "h2 { margin:0 0 12px; font-size:1.05rem; text-transform:uppercase; color:#64748b; }"
    )
    sections.append(
        ".summary-grid { display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:8px; }"
    )
    sections.append(".metric { border:1px solid #e5eaf0; border-radius:6px; padding:12px; }")
    sections.append(".metric b { display:block; font-size:1.5rem; }")
    sections.append(".metric span { color:#64748b; font-size:.72rem; text-transform:uppercase; }")
    sections.append(
        ".table-scroll { overflow-x:auto; border:1px solid #dce2ea; border-radius:6px; }"
    )
    sections.append(".result-table { margin-top:0; border:0; }")
    sections.append(".result-table th:first-child, .result-table td:first-child { border-left:0; }")
    sections.append(".result-table th:last-child, .result-table td:last-child { border-right:0; }")
    sections.append(".result-table tbody tr:last-child td { border-bottom:0; }")
    sections.append(".exception-rule-table th:last-child { text-align:right; }")
    sections.append(
        ".exception-table-sort { border:0; background:transparent; color:inherit; cursor:pointer; "
        "font:inherit; font-weight:700; text-transform:uppercase; } "
        ".exception-table-filter { display:block; width:100%; margin-top:5px; }"
    )
    sections.append(
        ".number-cell { text-align:right; font-weight:700; font-variant-numeric:tabular-nums; }"
    )
    sections.append(
        ".rule-id { display:inline-block; margin-right:8px; padding:2px 7px; "
        "border:1px solid #dce2ea; border-radius:999px; color:#64748b; "
        "background:#eef2f6; font-family:ui-monospace,monospace; font-size:.72rem; }"
    )
    sections.append(".comparison-columns { margin-bottom:12px; }")
    sections.append(
        ".comparison-columns-label { margin:0 0 4px; color:#64748b; font-size:.8rem; "
        "font-weight:600; text-transform:uppercase; letter-spacing:.03em; }"
    )
    sections.append(
        ".chip-list { display:flex; flex-wrap:wrap; gap:8px; "
        "list-style:none; margin:0; padding:0; }"
    )
    sections.append(
        ".tag { display:inline-block; padding:2px 8px; border:1px solid #dce2ea; "
        "border-radius:999px; background:#eef2f6; font-size:.75rem; font-weight:500; }"
    )
    sections.append(".group-stats-panel { margin-top:8px; }")
    sections.append(
        ".group-stats-row { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); "
        "gap:8px; margin-bottom:8px; align-items:stretch; }"
    )
    sections.append(".group-stats-row--1 { grid-template-columns:minmax(0,1fr); }")
    sections.append(".group-stats-row--2 { grid-template-columns:repeat(2,minmax(0,1fr)); }")
    sections.append(".group-stats-row--3 { grid-template-columns:repeat(3,minmax(0,1fr)); }")
    sections.append(
        ".group-stats-card { display:flex; flex-direction:column; min-width:0; "
        "background:#fff; border:1px solid #dce2ea; border-radius:8px; padding:16px; }"
    )
    sections.append(".group-stats-summary { cursor:pointer; font-weight:600; }")
    sections.append(".group-stats-column { display:block; font-family:ui-monospace,monospace; }")
    sections.append(
        ".group-stats-count { display:block; margin-top:2px; color:#64748b; "
        "font-size:.85rem; font-weight:400; }"
    )
    sections.append(".group-stats-scroll { max-height:192px; overflow-y:auto; }")
    sections.append(".group-stats-table { margin-top:8px; }")
    sections.append(
        "table { border-collapse:collapse; width:100%; margin-top:12px; font-size:.85rem; }"
    )
    sections.append(
        "th, td { border:1px solid #dce2ea; padding:8px 10px; "
        "text-align:left; vertical-align:top; }"
    )
    sections.append(
        "th { background:#eef2f6; color:#64748b; font-size:.72rem; text-transform:uppercase; }"
    )
    sections.append("code { white-space:pre-wrap; }")
    sections.append(
        "@media (max-width:900px) { .group-stats-row { grid-template-columns:"
        "repeat(3,minmax(0,1fr)); } .group-stats-row--1 { grid-template-columns:1fr; } "
        ".group-stats-row--2 { grid-template-columns:repeat(2,minmax(0,1fr)); } }"
    )
    sections.append(
        "@media (max-width:680px) { .summary-grid { grid-template-columns:1fr 1fr; } "
        ".group-stats-row { grid-template-columns:repeat(2,minmax(0,1fr)); } "
        ".group-stats-row--1 { grid-template-columns:1fr; } }"
    )
    sections.append("@media (max-width:520px) { .group-stats-row { grid-template-columns:1fr; } }")
    sections.append("</style>")
    sections.append("</head><body><main>")

    sections.append("<div class='report-header'>")
    sections.append(f"<h1>{_escape_html(report_name)}</h1>")
    formatted_time = _format_created_at(created_at)
    if formatted_time:
        sections.append(f"<span class='run-time'>Ran on {_escape_html(formatted_time)}</span>")
    sections.append("</div>")

    sections.append("<section class='card' id='overall'>")
    sections.append("<h2>Overall result</h2>")
    rows_a = comparison.get("total_rows_a", 0)
    rows_b = comparison.get("total_rows_b", 0)
    changes = comparison.get("rows_with_changes", 0)
    attr_changes = comparison.get("total_attribute_changes", 0)
    violation_rows = validation.get(
        "distinct_violating_rows", validation.get("total_violations", 0)
    )
    violation_attrs = validation.get(
        "distinct_violating_attributes", validation.get("total_violations", 0)
    )
    targets = result.get("target_columns") or []
    filters = result.get("filters_applied") or []
    filter_text = (
        "; ".join(_format_filter_row(f) for f in filters) if filters else "No filtering applied"
    )
    sections.append(
        "<p class='section-logic'>"
        f"Comparison across {len(targets)} target columns."
        f"<br>Filtering: {_escape_html(filter_text)}"
        "</p>"
    )
    sections.append("<div class='summary-grid'>")
    sections.append(_metric("Books after filters", rows_a + rows_b))
    sections.append(_metric("Books with rule exception", violation_rows))
    sections.append(_metric("Attributes with rule exception", violation_attrs))
    sections.append(_metric("Books with changes", changes))
    sections.append(_metric("Attributes changed", attr_changes))
    sections.append("</div>")

    grp = result.get("group_statistics") or {}
    overall_grp = grp.get("overall") or []
    sections.append(_render_group_section("Aggregation by grouping columns", overall_grp))
    sections.append("</section>")

    new_book_count = comparison.get("new_book_count", 0)
    extra_display = result.get("extra_column_display") or {}
    selected_extra_columns = list(result.get("exception_columns") or [])
    new_books_grp = grp.get("new_books") or []
    if new_book_count > 0:
        sections.append("<section class='card' id='new-books'>")
        sections.append("<h2>New Books</h2>")
        sections.append(_metric("new books found", new_book_count))
        if new_books_grp:
            sections.append(_render_group_section("New Books aggregation", new_books_grp))
        nb_details = comparison.get("new_book_details") or []
        if nb_details:
            sections.append("<table>")
            sections.append("<thead><tr>")
            for k in key_columns:
                sections.append(f"<th>{_escape_html(k)}</th>")
            if extra_display.get("new_books_html_report"):
                for column in selected_extra_columns:
                    sections.append(f"<th>{_escape_html(column)}</th>")
            sections.append("</tr></thead><tbody>")
            for nb in nb_details:
                sections.append("<tr>")
                for k in key_columns:
                    val = nb.get("key_columns", {}).get(k, "")
                    sections.append(f"<td>{_escape_html(str(val)) if val is not None else ''}</td>")
                if extra_display.get("new_books_html_report"):
                    values = nb.get("extra_values") or {}
                    for column in selected_extra_columns:
                        sections.append(f"<td>{_escape_html(values.get(column, ''))}</td>")
                sections.append("</tr>")
            sections.append("</tbody></table>")
        sections.append("</section>")

    row_details = comparison.get("row_details") or []

    comparison_sections = result.get("comparison_sections") or []
    if comparison_sections:
        for section in comparison_sections:
            section_name = section.get("name") or "Attribute Comparing Section"
            section_columns = set(section.get("columns") or [])
            section_rows = [
                (row, change)
                for row in row_details
                for change in row.get("attribute_changes", [])
                if change.get("column") in section_columns
            ]
            sections.append("<section class='card'>")
            sections.append(f"<h2>{_escape_html(section_name)}</h2>")
            sections.append(
                "<p class='section-logic'><code>In Baseline ≠ In Comparison</code> "
                f"— {len(section_columns)} column{'s' if len(section_columns) != 1 else ''}</p>"
            )
            if section_rows:
                sections.append("<table>")
                sections.append(_detail_header(key_columns, extra_columns=selected_extra_columns if extra_display.get("overall_html_report") else None))
                for row, change in section_rows:
                    sections.append("<tr>")
                    sections.append(_identity_cells(key_columns, row.get("key_columns", {}), row.get("row_index", "")))
                    if extra_display.get("overall_html_report"):
                        values = row.get("extra_values") or {}
                        for column in selected_extra_columns:
                            sections.append(f"<td>{_escape_html(values.get(column, ''))}</td>")
                    sections.append(f"<td>{_escape_html(change.get('column', ''))}</td>")
                    sections.append(f"<td>{_escape_html(change.get('file_a_value', ''))}</td>")
                    sections.append(f"<td>{_escape_html(change.get('file_b_value', ''))}</td>")
                    sections.append("</tr>")
                sections.append("</table>")
            else:
                sections.append(f"<p>No books with {_escape_html(section_name)}</p>")
            sections.append("</section>")
    else:
        sections.append("<section class='card' id='changes'>")
        sections.append("<h2>Attribute changes</h2>")
        sections.append(
            "<p class='section-logic'><code>In Baseline ≠ In Comparison</code> "
            "on shared target columns.</p>"
        )
        sections.append(_render_comparing_columns(result))
        attr_changes_grp = grp.get("attribute_changes") or []
        sections.append(_render_group_section("Attribute change aggregation", attr_changes_grp))
        if row_details:
            sections.append("<table>")
            sections.append(_detail_header(key_columns, extra_columns=selected_extra_columns if extra_display.get("overall_html_report") else None))
            for row in row_details:
                key_values = row.get("key_columns", {})
                for change in row.get("attribute_changes", []):
                    sections.append("<tr>")
                    sections.append(_identity_cells(key_columns, key_values, row.get("row_index", "")))
                    if extra_display.get("overall_html_report"):
                        values = row.get("extra_values") or {}
                        for column in selected_extra_columns:
                            sections.append(f"<td>{_escape_html(values.get(column, ''))}</td>")
                    sections.append(f"<td>{_escape_html(change.get('column', ''))}</td>")
                    sections.append(f"<td>{_escape_html(change.get('file_a_value', ''))}</td>")
                    sections.append(f"<td>{_escape_html(change.get('file_b_value', ''))}</td>")
                    sections.append("</tr>")
            sections.append("</table>")
        else:
            sections.append("<p>No detail rows.</p>")
        sections.append("</section>")

    sections.append(_render_exception_rule_summary(validation))

    violations_by_rule = validation.get("violations_by_rule") or {}
    rule_summaries = validation.get("rule_summaries") or {}
    rule_ids = list(dict.fromkeys([*violations_by_rule, *rule_summaries]))
    for rule_id in rule_ids:
        violations = violations_by_rule.get(rule_id) or []
        summary = rule_summaries.get(rule_id) or {}
        sample = violations[0] if violations else {}
        rule_name = summary.get("name") or sample.get("rule_name") or rule_id
        logic = (
            summary.get("logic")
            or sample.get("rule_logic")
            or "Rule details unavailable for this older run"
        )
        row_count = validation.get("violating_rows_by_rule", {}).get(rule_id, len(violations))
        attribute_count = validation.get("violating_attributes_by_rule", {}).get(
            rule_id, len(violations)
        )
        sections.append(f"<section class='card' id='rule-{_escape_html(rule_id)}'>")
        sections.append(f"<h2>{_escape_html(rule_id)} — {_escape_html(rule_name)}</h2>")
        if summary.get("description"):
            sections.append(
                "<p class='rule-description'>"
                f"{_escape_html(summary['description'])}</p>"
            )
        if summary.get("condition"):
            sections.append(
                "<p class='section-logic'>Condition: "
                f"<code>{_escape_html(summary['condition'])}</code></p>"
            )
        if summary.get("condition_grouping"):
            sections.append(
                "<p class='section-logic'>Grouping: "
                f"<code>{_escape_html(summary['condition_grouping'])}</code></p>"
            )
        sections.append(
            "<p class='section-logic'>Expectation: "
            f"<code>{_escape_html(_humanize_rule_logic(logic))}</code></p>"
        )
        sections.append("<div class='summary-grid'>")
        sections.append(_metric("Books with Exception", row_count))
        sections.append(_metric("Attributes with exception", attribute_count))
        sections.append("</div>")
        rule_grp = (grp.get("validation_rules") or {}).get(rule_id, [])
        sections.append(_render_group_section("Aggregation by rule grouping columns", rule_grp))
        if violations:
            extra_columns = list(
                dict.fromkeys(
                    column
                    for violation in violations
                    for column in (violation.get("extra_values") or {})
                )
            )
            hide_comparison = bool(summary.get("hide_comparison", False))
            sections.append("<table>")
            sections.append(
                _detail_header(
                    key_columns,
                    extra_columns=extra_columns,
                    hide_comparison=hide_comparison,
                )
            )
            for violation in violations:
                sections.append("<tr>")
                sections.append(
                    _identity_cells(
                        key_columns,
                        violation.get("key_columns", {}),
                        violation.get("row_index", ""),
                    )
                )
                extra_values = violation.get("extra_values") or {}
                for column in extra_columns:
                    sections.append(f"<td>{_escape_html(extra_values.get(column, '—'))}</td>")
                if not hide_comparison:
                    sections.append(
                        f"<td>{_escape_html(violation.get('violating_column', rule_id))}</td>"
                    )
                    sections.append(
                        f"<td>{_escape_html(violation.get('comparison_value', ''))}</td>"
                    )
                    sections.append(
                        f"<td>{_escape_html(violation.get('violating_value', ''))}</td>"
                    )
                sections.append("</tr>")
            sections.append("</table>")
        else:
            sections.append("<p>Nil exception detected under current rule.</p>")
        sections.append("</section>")

    sections.append(_render_exception_table(result, validation, key_columns))
    sections.append(_EXCEPTION_TABLE_INTERACTIONS)
    sections.append("</main></body></html>")
    return "\n".join(sections)


def _format_key(key_columns: dict[str, Any]) -> str:
    return ", ".join(f"{k}={v}" for k, v in key_columns.items())


def export_csv(result: dict[str, Any], report_name: str) -> str:
    output = io.StringIO()
    writer = csv.writer(output)

    comparison = result.get("comparison", {})
    validation = result.get("validation", {})

    writer.writerow(["Section", "Metric", "Value"])
    _write_summary_row(writer, "Total Rows A", comparison.get("total_rows_a", 0))
    _write_summary_row(writer, "Total Rows B", comparison.get("total_rows_b", 0))
    _write_summary_row(writer, "Rows with Changes", comparison.get("rows_with_changes", 0))
    attr_chg = comparison.get("total_attribute_changes", 0)
    _write_summary_row(writer, "Total Attribute Changes", attr_chg)
    _write_summary_row(writer, "Total Violations", validation.get("total_violations", 0))
    writer.writerow([])

    row_details = comparison.get("row_details", [])
    if row_details:
        writer.writerow(["Changes", "Row", "Key", "Column", "A", "B"])
        for row in row_details:
            key_str = _format_key(row.get("key_columns", {}))
            for change in row.get("attribute_changes", []):
                writer.writerow(
                    [
                        "Change",
                        _sanitize_csv_value(row.get("row_index", "")),
                        _sanitize_csv_value(key_str),
                        _sanitize_csv_value(change.get("column", "")),
                        _sanitize_csv_value(change.get("file_a_value", "")),
                        _sanitize_csv_value(change.get("file_b_value", "")),
                    ]
                )

    violations_by_rule = validation.get("violations_by_rule", {})
    if violations_by_rule:
        writer.writerow([])
        writer.writerow(["Violations", "Row", "Rule", "Key", "Details"])
        for rule_id, violations in violations_by_rule.items():
            for v in violations:
                key_str = _format_key(v.get("key_columns", {}))
                writer.writerow(
                    [
                        "Violation",
                        _sanitize_csv_value(v.get("row_index", "")),
                        _sanitize_csv_value(rule_id),
                        _sanitize_csv_value(key_str),
                        _sanitize_csv_value(v.get("details", "")),
                    ]
                )

    return output.getvalue()


def export_excel(result: dict[str, Any], report_name: str) -> bytes:
    """Build the multi-sheet result workbook used by the Excel export."""
    workbook = Workbook()
    overall = workbook.active
    overall.title = "Overall"
    comparison = result.get("comparison") or {}
    validation = result.get("validation") or {}
    group_statistics = result.get("group_statistics") or {}

    overall.append(["Overall Results"])
    overall.append(["Report name", _excel_value(report_name)])
    overall.append(
        ["Books after filters", comparison.get("total_rows_a", 0) + comparison.get("total_rows_b", 0)]
    )
    overall.append(
        [
            "Exception records",
            validation.get("distinct_violating_rows", validation.get("total_violations", 0)),
        ]
    )
    overall.append(
        [
            "Attributes with exception",
            validation.get("distinct_violating_attributes", validation.get("total_violations", 0)),
        ]
    )
    overall.append(["Rows changed", comparison.get("rows_with_changes", 0)])
    overall.append(["Attributes changed", comparison.get("total_attribute_changes", 0)])
    overall.append(["New Books", comparison.get("new_book_count", 0)])
    overall.append([])
    overall.append(["Filtering information"])
    filters = result.get("filters_applied") or []
    if filters:
        for filter_row in filters:
            overall.append([_excel_value(_format_filter_row(filter_row))])
    else:
        overall.append(["No filtering applied"])

    aggregation_start_row = overall.max_row + 2
    aggregation_bottom_row = aggregation_start_row
    overall_aggregations = group_statistics.get("overall") or []
    if overall_aggregations and 2 + (len(overall_aggregations) - 1) * 3 > _EXCEL_MAX_COLUMNS:
        raise ValueError("Overall aggregations exceed Excel's 16,384-column worksheet limit.")
    for index, stat in enumerate(overall_aggregations):
        start_column = 1 + index * 3
        overall.cell(
            aggregation_start_row,
            start_column,
            f"Aggregation: {stat.get('column', '?')}",
        )
        overall.cell(aggregation_start_row + 1, start_column, "Value")
        overall.cell(aggregation_start_row + 1, start_column + 1, "Exception records")
        for row_offset, stat_row in enumerate(stat.get("rows", []), start=2):
            overall.cell(
                aggregation_start_row + row_offset,
                start_column,
                _excel_value(stat_row.get("value", "")),
            )
            overall.cell(
                aggregation_start_row + row_offset,
                start_column + 1,
                stat_row.get("unique_count", 0),
            )
            aggregation_bottom_row = max(
                aggregation_bottom_row,
                aggregation_start_row + row_offset,
            )

    next_row = aggregation_bottom_row + 2
    overall.cell(next_row, 1, "Exception Rule Summary")
    next_row += 1
    overall.cell(next_row, 1, "Rule name")
    overall.cell(next_row, 2, "Exception records")
    summaries = validation.get("rule_summaries") or {}
    violations_by_rule = validation.get("violations_by_rule") or {}
    rule_ids = list(dict.fromkeys([*violations_by_rule, *summaries]))
    row_counts = validation.get("violating_rows_by_rule") or {}
    for rule_id in rule_ids:
        next_row += 1
        violations = violations_by_rule.get(rule_id) or []
        sample = violations[0] if violations else {}
        summary = summaries.get(rule_id) or {}
        rule_name = summary.get("name") or sample.get("rule_name") or rule_id
        overall.cell(next_row, 1, _excel_value(rule_name))
        overall.cell(next_row, 2, row_counts.get(rule_id, len(violations)))

    rule_summary_sheet = workbook.create_sheet("Rule Summary")
    rule_summary_sheet["A1"] = "Rule Summary"
    _append_row(
        rule_summary_sheet,
        3,
        ["Rule index", "Rule name", "Description", "Exception records"],
    )
    for row_number, rule_id in enumerate(rule_ids, start=4):
        violations = violations_by_rule.get(rule_id) or []
        sample = violations[0] if violations else {}
        summary = summaries.get(rule_id) or {}
        _append_row(
            rule_summary_sheet,
            row_number,
            [
                rule_id,
                summary.get("name") or sample.get("rule_name") or rule_id,
                summary.get("description", ""),
                row_counts.get(rule_id, len(violations)),
            ],
        )

    changes_sheet = workbook.create_sheet("Attribute Changes")
    changes_sheet["A1"] = "Attribute Changes"
    comparing_columns = result.get("target_columns") or result.get("common_columns") or []
    extra_display = result.get("extra_column_display") or {}
    selected_extra_columns = list(result.get("exception_columns") or [])
    changes_sheet["A2"] = "Comparing columns"
    changes_sheet["B2"] = _excel_value(", ".join(str(column) for column in comparing_columns))
    change_headers = [
        *(result.get("key_columns") or ["Row"]),
        "Column",
        "In Baseline",
        "In Comparison",
        *(
            [f"{column}(Latest Value)" for column in selected_extra_columns]
            if extra_display.get("overall_excel_report")
            else []
        ),
    ]
    change_count = sum(
        len(detail.get("attribute_changes") or []) for detail in comparison.get("row_details") or []
    )
    if change_count + 4 > _EXCEL_MAX_ROWS:
        raise ValueError("Attribute Changes exceeds Excel's 1,048,576-row worksheet limit.")
    _append_row(changes_sheet, 4, change_headers)
    change_row = 5
    key_columns = list(result.get("key_columns") or [])
    for detail in comparison.get("row_details") or []:
        identity = _excel_identity(
            key_columns, detail.get("key_columns") or {}, detail.get("row_index", "")
        )
        for change in detail.get("attribute_changes") or []:
            _append_row(
                changes_sheet,
                change_row,
                [
                    *identity,
                    change.get("column", ""),
                    change.get("file_a_value", ""),
                    change.get("file_b_value", ""),
                    *([(detail.get("extra_values") or {}).get(column, "") for column in selected_extra_columns] if extra_display.get("overall_excel_report") else []),
                ],
            )
            change_row += 1

    new_book_count = comparison.get("new_book_count", 0)
    if new_book_count > 0:
        nb_sheet = workbook.create_sheet("New Books")
        nb_sheet["A1"] = "New Books"
        nb_sheet["A2"] = "Books only in comparison file (not in baseline)"
        nb_sheet["B2"] = new_book_count
        nb_details = comparison.get("new_book_details") or []
        nb_headers = [*key_columns, *(selected_extra_columns if extra_display.get("new_books_excel_report") else [])]
        if len(nb_details) + 4 > _EXCEL_MAX_ROWS:
            raise ValueError("New Books exceeds Excel's 1,048,576-row worksheet limit.")
        _append_row(nb_sheet, 4, nb_headers)
        nb_row_num = 5
        for nb in nb_details:
            identity = _excel_identity(
                key_columns, nb.get("key_columns") or {}, nb.get("row_index", "")
            )
            values = [*identity]
            if extra_display.get("new_books_excel_report"):
                extras = nb.get("extra_values") or {}
                values.extend(extras.get(column, "") for column in selected_extra_columns)
            _append_row(nb_sheet, nb_row_num, values)
            nb_row_num += 1

    comparison_sections = result.get("comparison_sections") or []
    if comparison_sections:
        sections_sheet = workbook.create_sheet("Attribute Comparing Sections")
        next_row = 1
        for section in comparison_sections:
            section_columns = list(section.get("columns") or [])
            if not section_columns:
                continue

            section_headers = [
                *(key_columns or ["Row"]),
                "Column",
                "In Baseline",
                "In Comparison",
            ]

            section_name = section.get("name") or "Attribute Comparing Section"
            sections_sheet.cell(next_row, 1, _excel_value(section_name))
            sections_sheet.cell(next_row + 1, 1, "Comparing columns")
            sections_sheet.cell(
                next_row + 1,
                2,
                _excel_value(", ".join(str(column) for column in section_columns)),
            )
            for column_offset, header in enumerate(section_headers):
                sections_sheet.cell(next_row + 3, column_offset + 1, header)

            section_row = next_row + 4
            record_count = 0
            for detail in comparison.get("row_details") or []:
                identity = _excel_identity(
                    key_columns, detail.get("key_columns") or {}, detail.get("row_index", "")
                )
                for change in detail.get("attribute_changes") or []:
                    if change.get("column") not in section_columns:
                        continue
                    if section_row > _EXCEL_MAX_ROWS:
                        raise ValueError(
                            "An Attribute Comparing Section exceeds Excel's 1,048,576-row worksheet limit."
                        )
                    _append_row(
                        sections_sheet,
                        section_row,
                        [
                            *identity,
                            change.get("column", ""),
                            change.get("file_a_value", ""),
                            change.get("file_b_value", ""),
                        ],
                    )
                    section_row += 1
                    record_count += 1

            if record_count == 0:
                sections_sheet.cell(section_row, 1, "There is 0 record for this table.")
                section_row += 1

            # Keep one blank row between adjacent section tables.
            next_row = section_row + 1

    for rule_id in rule_ids:
        violations = violations_by_rule.get(rule_id) or []
        summary = summaries.get(rule_id) or {}
        sample = violations[0] if violations else {}
        rule_name = summary.get("name") or sample.get("rule_name") or rule_id
        sheet = workbook.create_sheet(_safe_sheet_name(rule_id, workbook.sheetnames))
        sheet["A1"] = _excel_value(f"{rule_id} - {rule_name}")
        sheet["A2"] = "Condition:"
        sheet["B2"] = _excel_value(summary.get("condition", ""))
        sheet["A3"] = "Grouping:"
        sheet["B3"] = _excel_value(summary.get("condition_grouping", ""))
        expectation = summary.get("logic") or sample.get("rule_logic") or ""
        sheet["A4"] = "Expectation:"
        sheet["B4"] = _excel_value(_humanize_rule_logic(expectation))

        extra_columns = list(
            dict.fromkeys(
                column
                for violation in violations
                for column in (violation.get("extra_values") or {})
            )
        )
        hide_comparison = bool(summary.get("hide_comparison", False))
        if len(violations) + 7 > _EXCEL_MAX_ROWS:
            raise ValueError(f"Rule {rule_id} exceeds Excel's 1,048,576-row worksheet limit.")
        headers = [*(key_columns or ["Row"]), *extra_columns]
        if not hide_comparison:
            headers.extend(["Column", "In Baseline", "In Comparison"])
        _append_row(sheet, 7, headers)
        row_number = 8
        for violation in violations:
            values = _excel_identity(
                key_columns,
                violation.get("key_columns") or {},
                violation.get("row_index", ""),
            )
            extra_values = violation.get("extra_values") or {}
            values.extend(extra_values.get(column, "—") for column in extra_columns)
            if not hide_comparison:
                values.extend(
                    [
                        violation.get("violating_column", rule_id),
                        violation.get("comparison_value", ""),
                        violation.get("violating_value", ""),
                    ]
                )
            _append_row(sheet, row_number, values)
            row_number += 1

    agg_columns = list(result.get("aggregation_columns") or [])
    agg_labels = result.get("aggregation_column_labels") or {}
    exception_cols = _exception_table_columns(result, violations_by_rule)
    exception_col_set = set(exception_cols)

    exc_sheet = workbook.create_sheet("Exception Table")
    exc_sheet["A1"] = "Exception Table"
    exc_headers = [
        *(key_columns or ["Row"]),
        "Rule Index",
        *(
            agg_labels.get(col, col) for col in agg_columns
        ),
        *exception_cols,
    ]
    _append_row(exc_sheet, 3, exc_headers)
    exc_row = 4

    for rule_id in rule_ids:
        violations = violations_by_rule.get(rule_id) or []
        for violation in violations:
            grouping = violation.get("grouping_values") or {}
            extra_vals = violation.get("extra_values") or {}
            filtered_extra = {k: v for k, v in extra_vals.items() if k in exception_col_set}
            if exc_row > _EXCEL_MAX_ROWS:
                raise ValueError("Exception Table exceeds Excel's 1,048,576-row worksheet limit.")
            _append_row(
                exc_sheet,
                exc_row,
                [
                    *_excel_identity(
                        key_columns,
                        violation.get("key_columns") or {},
                        violation.get("row_index", ""),
                    ),
                    _excel_value(rule_id),
                    *(
                        _excel_value(grouping.get(col, "")) for col in agg_columns
                    ),
                    *(
                        _excel_value(filtered_extra.get(col, "")) for col in exception_cols
                    ),
                ],
            )
            exc_row += 1

    for sheet in workbook.worksheets:
        _style_excel_sheet(sheet)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_row(sheet: Any, row: int, values: list[Any]) -> None:
    for column, value in enumerate(values, start=1):
        sheet.cell(row, column, _excel_value(value))


def _excel_identity(key_columns: list[str], values: dict[str, Any], row_index: Any) -> list[Any]:
    if not key_columns:
        return [row_index]
    return [values.get(column, "—") for column in key_columns]


def _excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return _sanitize_csv_value(value)
    return value


def _safe_sheet_name(name: str, existing: list[str]) -> str:
    cleaned = "".join("_" if char in r"[]:*?/\\" else char for char in name)[:31] or "Rule"
    candidate = cleaned
    index = 2
    while candidate in existing:
        suffix = f"_{index}"
        candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
        index += 1
    return candidate


def _style_excel_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, size=14)
            if isinstance(cell.value, str) and (
                cell.value
                in {
                    "Value",
                    "Exception records",
                    "Rule name",
                    "Column",
                    "In Baseline",
                    "In Comparison",
                    "Condition:",
                    "Grouping:",
                    "Expectation:",
                }
                or cell.value in {"Filtering information", "Exception Rule Summary"}
                or "aggregation:" in cell.value
            ):
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for column_cells in sheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=0)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_length + 2, 12), 50
        )
    sheet.freeze_panes = "A2"


def _write_summary_row(writer: Any, metric: str, value: Any) -> None:
    writer.writerow(["Summary", metric, _sanitize_csv_value(value)])
